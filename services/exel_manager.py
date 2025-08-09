import pandas as pd
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional, Tuple
import json

class ExcelManager:
    
    def __init__(self, sheet_name: str = "Form"):
        self.header_row = 2
        self.order_name_cell = "B1"
        self.sheet_name = sheet_name
        self.default_formats = {
            'border': {'border': 1},
            'header': {'border': 1, 'bold': True, 'bg_color': '#D7E4BC'},
            'order_name': {'border': 1, 'bg_color': '#FFE699', 'bold': True}
        }
    
    def create_form(self,
                   filename: str,
                   all_columns: List[Dict[str, Any]],
                   num_rows: int = 10,
                   order_name_position: str = "A1",
                   formats: Optional[Dict[str, Dict]] = None) -> None:

        if formats:
            self.default_formats.update(formats)
        
        # Gabungkan header dan specs
        columns = [item['name'] for item in all_columns]
        
        # Buat data kosong
        data = []
        for i in range(1, num_rows + 1):
            row = [i] + ['' for _ in range(len(columns) - 1)]
            data.append(row)
        
        df = pd.DataFrame(data, columns=columns)
        
        # Tulis ke Excel
        with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name=self.sheet_name, startrow=self.header_row, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[self.sheet_name]
            
            # Terapkan format
            border_format = workbook.add_format(self.default_formats['border'])
            header_format = workbook.add_format(self.default_formats['header'])
            order_name_format = workbook.add_format(self.default_formats['order_name'])
            
            # Tulis Order Name
            order_row, order_col = self._parse_cell_position(order_name_position)
            worksheet.write(order_row, order_col, "Order Name:", order_name_format)
            worksheet.write(order_row, order_col + 1, "", order_name_format)
            
            # Header
            for col in range(len(columns)):
                worksheet.write(self.header_row, col, columns[col], header_format)
                worksheet.write_comment(self.header_row, col, json.dumps(all_columns[col]))
            
            num_index = 1
            # Border untuk baris data
            for row in range(self.header_row + 1, self.header_row + 1 + len(df)):
                for col in range(len(columns)):
                    if columns[col] == 'No':
                        value = num_index
                    else:
                        value = ""
                    worksheet.write(row, col, value, border_format)
                num_index += 1
            
            # Set lebar kolom
            for i, col in enumerate(columns):
                max_len = max(len(str(col)), 15)
                worksheet.set_column(i, i, max_len + 2)
    
    def read_form(self, file_bytes, filter_column: Optional[str] = None) -> Tuple[str, Dict[str, Any]]:

        try:
            # Baca data utama
            df = pd.read_excel(file_bytes, sheet_name=self.sheet_name, header=self.header_row)
            
            order_name = None
            form_data_raw = df
            
            # Baca nama order
            wb = load_workbook(file_bytes)
            ws = wb[self.sheet_name]
            order_name = ws[self.order_name_cell].value
            order_name = order_name
            
            # Filter data jika ada kolom filter
            # if filter_column and filter_column in df.columns:
            #    form_data = df[df[filter_column].notna()]

            #ganti header dengan komentar
            header_comments = [
                cell.comment.text if cell.comment else cell.value
                for cell in ws[self.header_row + 1]
            ]
            df.columns = header_comments

            form_data = self.get_filled_rows(form_data_raw)
            
            return order_name, form_data
            
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def get_filled_rows(self, data: pd.DataFrame) -> List[Dict[str, Any]]:

        filled_rows = []
        
        for idx, row in data.iterrows():
            filled = row.dropna()
            if not filled.empty:
                filled_rows.append(dict(filled))
        
        return filled_rows
    
    """ def update_cell(self, filename: str, cell: str, value: Any, save: bool = True) -> None:
        # Update nilai
        
        try:
            wb = load_workbook(filename)
            ws = wb[self.sheet_name]
            ws[cell] = value
            
            if save:
                wb.save(filename)
                
        except Exception as e:
            raise Exception(f"Error updating cell: {str(e)}") """
    
    def _parse_cell_position(self, cell: str) -> tuple:
        # Parse posisi cell dari string ke koordinat row, col

        from openpyxl.utils import column_index_from_string
        
        col_str = ''.join([c for c in cell if c.isalpha()])
        row_str = ''.join([c for c in cell if c.isdigit()])
        
        col = column_index_from_string(col_str) - 1  # Convert to 0-indexed
        row = int(row_str) - 1  # Convert to 0-indexed
        
        return row, col